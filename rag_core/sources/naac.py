"""NAAC accreditation grades — the one quality signal the catalogue has none of.

WHY THIS EXISTS
`colleges.naac_grade` is empty on 38,698 of 38,700 rows. The column exists and
the detail crawl returns the key, but the client's own source DB never had the
value, so no amount of crawling their API produces it. Students ask for the NAAC
grade constantly; MIVI currently cannot answer at all.

NAAC publishes it itself, which makes this a registry read rather than a scrape:

    naac.gov.in/.../Institutions_accredited_by_NAAC_having_valid_accreditation
                    -as_on_14082025_1.xlsx

    sheet                            rows    columns
    Universities                      498    HEI Name, Track-Id, Aishe-Id,
    Colleges                        7,567    Address, Current Cycle Number,
    Transition Autonomous Colleges    291    Current CGPA, Current Grade,
                                             Date Of Declaration

8,356 institutions, each with a grade AND the CGPA behind it AND the date it was
declared. Two things make it trustworthy where a scraped grade would not be:

- It is the list of accreditation that was VALID on the file's date, so a lapsed
  accreditation is absent rather than stale-but-present.
- The grade distribution runs A++ (84) through C (12) across universities, so it
  is not a "top colleges" selection that would make absence look like failure.

FRESHNESS IS RENDERED, NOT ASSUMED
The file is dated 14-08-2025 and each row carries its own declaration date. A
NAAC cycle runs about five years, so a grade declared in 2019 may have been
re-assessed since. Every fact therefore stores its declaration date and renders
with it, the same discipline the Fee Regulating Authority lines use. MIVI states
when the grade was declared and tells the student to confirm — it never presents
a grade as "current".

MATCHING
Reuses the hardened matcher in registries.py rather than a second one: the
family gate (an IIT only matches an IIT), token containment, the dotted-acronym
gate and the leading-brand gate all took real mistakes to arrive at, and a
parallel implementation would repeat them. The one thing added here is city
recovery: the sheet has an Address but no city column, so the address text is
scanned for a place name the catalogue already knows. Finding one lets the row
be scored against that city's colleges at the 0.72 floor instead of falling to
the 0.92 no-city floor.

Run:
    python -m rag_core.sources.naac --dry-run
    python -m rag_core.sources.naac
    python -m rag_core.sources.naac --report
"""
from __future__ import annotations

import argparse
import json
import re
import sys
from datetime import datetime, timezone
from pathlib import Path

import httpx

from .. import config  # noqa: F401 - loads .env before the backend reads it
from ..store.backend import get_backend
from .registries import (_index_by_city, _index_by_token, _load_catalogue,
                         _load_place_aliases, _norm, _place_key, ensure_tables,
                         match_one)

ROOT = Path(__file__).resolve().parents[2]
CACHE_DIR = ROOT / "data" / "raw" / "registries" / "naac"

# The file name carries its own as-on date, so a newer publication caches
# separately instead of overwriting the evidence for facts already stored.
XLSX_URL = ("http://www.naac.gov.in/images/docs/ACCREDITATION_STATUS/"
            "Institutions_accredited_by_NAAC_having_valid_accreditation"
            "-as_on_14082025_1.xlsx")
AS_ON = "2025-08-14"

VALID_GRADES = ("A++", "A+", "A", "B++", "B+", "B", "C")

# NAAC's own sheet names, mapped to the category recorded on each fact.
#
# The workbook's third sheet, "Transition Autonomous Colleges", is deliberately
# NOT read. Its columns are Sl. No. / HEI Name / State / Extended validity upto
# — it lists autonomous-status extensions, not accreditation, and carries no
# grade or CGPA. Its validity dates had already passed (2024-12-31, 2025-12-31)
# when this was written. Naming it here so nobody reads "290 rows skipped" as a
# parser bug and 'fixes' it into the grade table.
SHEETS = {
    "Universities": "university",
    "Colleges": "college",
}


def fetch_xlsx(*, refresh: bool = False) -> Path:
    CACHE_DIR.mkdir(parents=True, exist_ok=True)
    dest = CACHE_DIR / f"naac_valid_{AS_ON}.xlsx"
    if dest.exists() and not refresh:
        return dest
    print(f"[naac] downloading {XLSX_URL.rsplit('/', 1)[-1]}", file=sys.stderr)
    with httpx.Client(timeout=90, follow_redirects=True, verify=False) as c:
        r = c.get(XLSX_URL)
        r.raise_for_status()
    dest.write_bytes(r.content)
    print(f"[naac] cached {len(r.content) // 1024} KB -> {dest.name}",
          file=sys.stderr)
    return dest


def _cgpa(value) -> float | None:
    """A CGPA is on a 0-4 scale; anything else is a parse error, not a score."""
    try:
        f = float(str(value).strip())
    except (TypeError, ValueError):
        return None
    return f if 0.0 < f <= 4.0 else None


def _iso_date(value) -> str | None:
    """NAAC writes dd-mm-yyyy; Excel sometimes hands back a datetime."""
    if isinstance(value, datetime):
        return value.date().isoformat()
    s = str(value or "").strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            return datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    return None


def read_rows(path: Path) -> list[dict]:
    """Every graded institution across the three sheets."""
    from openpyxl import load_workbook

    wb = load_workbook(path, read_only=True, data_only=True)
    out: list[dict] = []
    for sheet, category in SHEETS.items():
        if sheet not in wb.sheetnames:
            print(f"[naac] sheet '{sheet}' missing from the workbook",
                  file=sys.stderr)
            continue
        ws = wb[sheet]
        header: dict[str, int] | None = None
        kept = skipped = 0
        for raw in ws.iter_rows(values_only=True):
            cells = ["" if c is None else str(c).strip() for c in raw]
            if not any(cells):
                continue
            if header is None:
                if any("hei name" in c.lower() for c in cells):
                    header = {c.strip().lower(): i for i, c in enumerate(cells) if c}
                continue

            def cell(*names, _cells=cells, _h=header):
                for nm in names:
                    i = _h.get(nm)
                    if i is not None and i < len(_cells):
                        return _cells[i]
                return ""

            name = " ".join(cell("hei name").split())
            grade = cell("current grade").upper().replace(" ", "")
            # No grade means nothing to tell a student. Recorded as skipped
            # rather than stored blank, so the count is visible.
            if not name or grade not in VALID_GRADES:
                skipped += 1
                continue
            out.append({
                "name": name,
                "category": category,
                "grade": grade,
                "cgpa": _cgpa(cell("current cgpa")),
                "cycle": cell("current cycle number") or None,
                "declared": _iso_date(cell("date of declaration")),
                "aishe_id": cell("aishe-id", "aishe id") or None,
                "track_id": cell("track-id", "track id") or None,
                "address": " ".join(cell("address").split()),
            })
            kept += 1
        print(f"[naac] {sheet}: {kept:,} graded, {skipped:,} without a usable grade",
              file=sys.stderr)
    return out


def _city_from_address(address: str, place_keys: set[str],
                       hei_name: str = "") -> str:
    """Recover a city the catalogue knows from NAAC's free-text address.

    The sheet has no city column, and city agreement is what decides whether a
    row is scored at the 0.72 floor or the 0.92 one, so a wrong city here is a
    real loss of safety rather than a cosmetic error. Two things it must not do:

    1. Read a place out of the INSTITUTION'S NAME. NAAC's Address cell routinely
       begins with the institution's own name ("Andhra University, karakachettu
       P.O ..."), and many Indian colleges are named after a place they are not
       in — "Nagpur Institute of Technology, Wardha". The name is removed before
       scanning, so only the address part can supply a city.

    2. Prefer a district over a city. `place_keys` must therefore be CITY keys
       only. registries.py now also indexes colleges by district, which put 135
       district-only names into that vocabulary; because those names are often
       longer than the town's, the longest-match rule below started choosing the
       district and matching against every college in it.

    Longest match still wins among cities, so "navi mumbai" beats "mumbai".
    """
    addr = _norm(address)
    name = _norm(hei_name)
    if name and addr.startswith(name):
        addr = addr[len(name):]
    elif name and len(name) > 12 and name in addr:
        addr = addr.replace(name, " ")
    hay = f" {addr} "
    best = ""
    for key in place_keys:
        if len(key) < 4 or len(key) <= len(best):
            continue
        if f" {key} " in hay:
            best = key
    return best


DDL = """
CREATE TABLE IF NOT EXISTS registry_fact (
    college_id TEXT NOT NULL, registry TEXT NOT NULL, ref_id TEXT,
    category TEXT, year TEXT, payload JSONB, source_url TEXT, detail_url TEXT,
    match_score REAL, match_note TEXT, fetched_at TIMESTAMPTZ,
    PRIMARY KEY (college_id, registry, category, year)
)
"""


def run(*, dry_run: bool = False, refresh: bool = False,
        limit: int | None = None) -> dict:
    ensure_tables()
    get_backend().execute(DDL.strip())

    rows = read_rows(fetch_xlsx(refresh=refresh))
    if limit:
        rows = rows[:limit]
    if not rows:
        print("[naac] nothing to match", file=sys.stderr)
        return {}

    colleges = _load_catalogue()
    aliases = _load_place_aliases()
    by_city = _index_by_city(colleges, aliases)
    by_token = _index_by_token(colleges)
    # CITY keys only — see _city_from_address. by_city also holds district keys,
    # and letting those into the address vocabulary made the longest-match rule
    # resolve to a district, whose bucket is every college in it.
    place_keys = {k for k in (_place_key(c.get("city"), aliases)
                              for c in colleges) if k}
    print(f"[naac] matching {len(rows):,} graded institutions against "
          f"{len(colleges):,} catalogue colleges", file=sys.stderr)

    be = get_backend()
    now = datetime.now(timezone.utc)
    stats = {"rows": len(rows), "matched": 0, "unmatched": 0, "no_city": 0,
             "contested": 0, "contested_colleges": 0}
    # college_id -> every NAAC row that claimed it. Contested colleges are
    # dropped wholesale; see the comment at the resolution step.
    claims: dict[str, list[tuple]] = {}
    misses: list[tuple] = []
    samples: list[str] = []

    for r in rows:
        city = _city_from_address(r["address"], place_keys, r["name"])
        if not city:
            stats["no_city"] += 1
        entry = {"name": r["name"], "city": city}
        college, score, why = match_one(entry, by_city, by_token, aliases)

        if college is None:
            stats["unmatched"] += 1
            misses.append(("naac", r["aishe_id"] or r["track_id"] or r["name"][:40],
                           r["category"], AS_ON, r["name"], city, None,
                           round(score, 3), why))
            continue

        stats["matched"] += 1
        payload = {"grade": r["grade"], "cgpa": r["cgpa"], "cycle": r["cycle"],
                   "declared": r["declared"], "aishe_id": r["aishe_id"],
                   "as_on": AS_ON, "naac_category": r["category"]}
        row = (college["college_id"], "naac",
               r["aishe_id"] or r["track_id"], "accreditation", AS_ON,
               json.dumps(payload), XLSX_URL, None,
               round(score, 3), why, now)
        # Claims are collected, not applied. Whether a claim is safe cannot be
        # judged one row at a time: it depends on whether anything ELSE claims
        # the same college. Resolved in one pass below.
        claims.setdefault(college["college_id"], []).append(
            (row, r["name"], r["grade"], city))
        if len(samples) < 14:
            samples.append(f"  {r['grade']:<4} {r['cgpa'] or '-':<5} "
                           f"{r['name'][:40]:40} -> {college['name'][:40]:40} "
                           f"({score:.2f})")

    # ---- resolve contested colleges by REFUSING them ----------------------
    # A college has exactly one NAAC grade. When two NAAC institutions both
    # match one college, at least one is a wrong grade on a real college — and
    # the higher score is NOT a safe tie-break. Measured on this workbook, 49
    # colleges were contested by 117 rows and 38 of those disagreed on the
    # grade, including cases where picking the top score is provably wrong:
    #
    #   ours: MRR College Of Pharmacy Nandigama
    #     0.750 A   NRI COLLEGE OF PHARMACY      <- wrong
    #     0.737 B++ SIMS COLLEGE OF PHARMACY     <- also wrong
    #   ours: Government First Grade College Hosadurga
    #     0.857 B+ (2.64) GOVERNMENT FIRST GRADE COLLEGE
    #     0.857 B  (2.33) GOVERNMENT FIRST GRADE COLLEGE   <- tied, disagreeing
    #
    # Karnataka alone has hundreds of colleges named exactly "Government First
    # Grade College", one per town, and a NAAC row that omits the town cannot be
    # assigned to any of them. Every name gate passes because the names are
    # identical; identical generic names are precisely the case where a score
    # carries no information.
    #
    # So a contested college gets NO grade. Telling a student their college is
    # B+ when it is C is worse than telling them nothing, and this is the one
    # decision where the safe answer is silence.
    writes: list[tuple] = []
    for cid, group in claims.items():
        if len(group) == 1:
            writes.append(group[0][0])
            continue
        stats["contested_colleges"] += 1
        stats["contested"] += len(group)
        grades = sorted({g[2] for g in group})
        for row, nm, grade, city in group:
            misses.append((
                "naac", f"contested:{row[2]}", "accreditation", AS_ON, nm, city,
                None, row[8],
                f"{len(group)} NAAC institutions matched the same college "
                f"({grades}) — refused rather than guessed"))

    for s in samples:
        print(s, file=sys.stderr)
    print(f"[naac] matched {stats['matched']:,}/{stats['rows']:,} "
          f"({100 * stats['matched'] / max(stats['rows'], 1):.1f}%), "
          f"{stats['unmatched']:,} refused, "
          f"{stats['no_city']:,} had no recoverable city"
          + ("  (DRY RUN — nothing written)" if dry_run else ""),
          file=sys.stderr)
    if stats["contested_colleges"]:
        print(f"[naac] REFUSED {stats['contested_colleges']:,} colleges claimed by "
              f"{stats['contested']:,} competing NAAC rows — a contested grade is "
              f"a coin flip, so none is stored. {len(writes):,} colleges will be "
              f"written", file=sys.stderr)

    if dry_run:
        return stats

    # Clean replace, not an upsert. Everything here is re-derived from the
    # workbook, and an upsert alone would strand grades that this run now
    # REFUSES — the contested colleges would silently keep the coin-flip grade a
    # previous run wrote, which is the exact failure the refusal exists to stop.
    removed = be.execute("DELETE FROM registry_fact WHERE registry = 'naac'")
    if removed:
        print(f"[naac] cleared {removed:,} grades from the previous run",
              file=sys.stderr)

    be.executemany(
        "INSERT INTO registry_fact (college_id, registry, ref_id, category, "
        "year, payload, source_url, detail_url, match_score, match_note, "
        "fetched_at) VALUES (?,?,?,?,?,?,?,?,?,?,?) "
        "ON CONFLICT (college_id, registry, category, year) DO UPDATE SET "
        "payload = EXCLUDED.payload, match_score = EXCLUDED.match_score, "
        "match_note = EXCLUDED.match_note, fetched_at = EXCLUDED.fetched_at",
        writes)
    try:
        be.executemany(
            "INSERT INTO registry_unmatched (registry, ref_id, category, year, "
            "name, city, state, best_score, note) VALUES (?,?,?,?,?,?,?,?,?) "
            "ON CONFLICT (registry, ref_id, category, year) DO UPDATE SET "
            "best_score = EXCLUDED.best_score, note = EXCLUDED.note", misses)
    except Exception as exc:  # noqa: BLE001 - the misses log is diagnostic only
        print(f"[naac] could not record misses ({str(exc)[:70]})", file=sys.stderr)

    n = be.one("SELECT COUNT(*) FROM registry_fact WHERE registry='naac'") or 0
    c = be.one("SELECT COUNT(DISTINCT college_id) FROM registry_fact "
               "WHERE registry='naac'") or 0
    print(f"[naac] stored {n:,} facts across {c:,} colleges", file=sys.stderr)
    stats["stored"] = n
    stats["colleges"] = c
    return stats


def report() -> None:
    be = get_backend()
    rows = be.q("SELECT payload->>'grade' AS g, COUNT(*) n FROM registry_fact "
                "WHERE registry='naac' GROUP BY 1 ORDER BY n DESC")
    total = sum(r["n"] for r in rows)
    print(f"NAAC grades on {total:,} colleges:")
    for r in rows:
        print(f"  {r['g'] or '?':<5} {r['n']:>6,}")


def main() -> None:
    ap = argparse.ArgumentParser(description=__doc__)
    ap.add_argument("--dry-run", action="store_true",
                    help="match and report, write nothing")
    ap.add_argument("--refresh", action="store_true",
                    help="re-download the workbook")
    ap.add_argument("--limit", type=int, help="only the first N graded rows")
    ap.add_argument("--report", action="store_true")
    a = ap.parse_args()
    if a.report:
        report()
        return
    run(dry_run=a.dry_run, refresh=a.refresh, limit=a.limit)


if __name__ == "__main__":
    main()
